import string
import json

from excel_writer import ExcelWriter, RowCollection
from entries import models as entry_models
from leads.models import Attachment
from openpyxl.styles import Font  # , Color
from django.db.models import Q
from entries.export_entries_docx import analysis_filter, xstr


def format_date(date):
    if date:
        try:
            return date.strftime('%d-%m-%Y')
        except:
            return date
    else:
        return None


def list_filter(_list, _filter, _value, key=None):
    try:
        _results = [_ele for _ele in _list
                    if _ele.get(_filter) and _ele.get(_filter) == _value]
        if len(_results) == 0:
            return None
        if key:
            return _results[0].get(key, '')
        return _results[0]
    except:
        return None


def get_analysis_data(elements, eID, element, eType, rows):
    """
    Try to retrive data from elements(JSON), if not found in template or
    error occurs, cell is filled with blank ''.
    """
    try:
        elementTemplate = list_filter(elements, 'id', eID)

        if eType == 'number-input':
            rows.add_value(element['value'])
        elif eType == 'date-input':
            rows.add_value(format_date(element['value']))

        elif eType == 'scale':
            rows.add_value(list_filter(elementTemplate['scaleValues'], 'id',
                                       element['value'], key='name'))

        elif eType == 'multiselect':
            multiselect_value = []
            for value in element.get('value', ''):
                option = list_filter(elementTemplate['options'], 'id', value,
                                     key='text')
                multiselect_value.append(option) if option else ''
            rows.add_value(', '.join(multiselect_value))

        elif eType == 'organigram':
            organigram_value = []
            for value in element.get('value', ''):
                option = list_filter(elementTemplate['nodes'], 'id', value,
                                     key='name')
                organigram_value.append(option) if option else ''
            rows.add_value(', '.join(organigram_value))

        elif eType == 'matrix1d':
            matrix_values = []
            selections = element.get('selections', [])
            for selection in selections:
                dimension = list_filter(elementTemplate.get('pillars', []),
                                        'id', selection.get('pillar'))
                sub_dimension = list_filter(dimension.get('subpillars', []),
                                            'id', selection.get('subpillar'),
                                            key='name') if dimension else ''
                matrix_values.append([
                    dimension.get('name') if dimension else '',
                    sub_dimension])
            rows.permute_and_add_list(matrix_values)

        elif eType == 'matrix2d':
            matrix_values = []
            selections = element.get('selections', [])
            for selection in selections:
                dimension = list_filter(elementTemplate.get('pillars', []),
                                        'id', selection.get('pillar'))
                sub_dimension = list_filter(dimension.get('subpillars', []),
                                            'id', selection.get('subpillar'),
                                            key='title') if dimension else ''
                sector = list_filter(elementTemplate.get('sectors'), 'id',
                                     selection.get('sector'))
                sub_sector = []
                if sector:
                    for _sub_sector in selection.get('subsectors', []):
                        sub_sector.append(
                                list_filter(sector.get('subsectors', []),
                                            'id', _sub_sector,
                                            key='title'))

                matrix_values.append([
                    dimension.get('title') if dimension else '',
                    sub_dimension, sector.get('title') if sector else '',
                    ', '.join(sub_sector)])
            rows.permute_and_add_list(matrix_values)

        elif eType == 'number2d':
            number_values = []
            matches = []
            numbers = element.get('numbers', [])

            for row in elementTemplate['rows']:
                same = True
                last = None
                for i, column in enumerate(elementTemplate['columns']):
                    n = next(x['value'] for x in numbers
                             if x['row'] == row['id'] and
                             x['column'] == column['id'])
                    number_values.append(n if n else '')

                    if same and i > 0 and last != n:
                        same = False
                    last = n
                matches.append('True' if same else 'False')

            number_values.extend(matches)
            rows.add_values(number_values)

    except Exception as e:
        if eType not in ['matrix1d', 'matrix2d', 'number2d']:
            rows.add_value('')
        elif eType == 'matrix1d':
            rows.add_values(['', ''])
        elif eType == 'matrix2d':
            rows.add_values(['', '', '', ''])
        elif eType == 'number2d':
            for row in elementTemplate['rows']:
                for column in elementTemplate['columns']:
                    rows.add_value('')
                rows.add_value('')


def export_xls(title, event_pk=None, information_pks=None):

    # Create a spreadsheet and get active workbook
    ew = ExcelWriter()
    ws = ew.get_active()
    ws.title = "Split Entries"
    wsg = ew.wb.create_sheet("Grouped Entries")

    # Create title row
    titles = [
        "Date of Lead Publication", "Date of Information", "Created By",
        "Date Imported", "URL", "Lead Title", "Source", "Excerpt",
        "Reliability",
        "Severity", "Demographic Groups", "Specific Needs Groups",
        "Affected Groups", "Pillar", "Subpillar", "Sector", "Subsector",
    ]

    if event_pk:
        countries = entry_models.Event.objects.get(pk=event_pk)\
            .countries.all().distinct()
    else:
        countries = entry_models.Country.objects.all().distinct()

    for country in countries:
        admin_levels = country.adminlevel_set.all().distinct()
        for admin_level in admin_levels:
            titles.append(admin_level.name)

    for i, t in enumerate(titles):
        ws.cell(row=1, column=i+1).value = t
        ws.cell(row=1, column=i+1).font = Font(bold=True)

        wsg.cell(row=1, column=i+1).value = t
        wsg.cell(row=1, column=i+1).font = Font(bold=True)

    ew.auto_fit_cells_in_row(1, ws)
    ew.auto_fit_cells_in_row(1, wsg)

    if event_pk:
        # Add each information in each entry belonging to this event
        informations = entry_models.EntryInformation.objects.filter(
                            entry__lead__event__pk=event_pk,
                            entry__template=None).distinct()
    else:
        # All information
        informations = entry_models.EntryInformation.objects.\
                                    filter(entry__template=None).distinct()

    if information_pks:
        informations = informations.filter(pk__in=information_pks)

    grouped_rows = []
    for i, info in enumerate(informations):
        try:
            rows = RowCollection(1)

            lead_url = info.entry.lead.url
            if Attachment.objects.filter(lead=info.entry.lead).count() > 0:
                lead_url = info.entry.lead.attachment.upload.url

            rows.add_values([
                format_date(info.entry.lead.published_at),
                format_date(info.date), info.entry.created_by,
                format_date(info.entry.created_at.date()),
                lead_url,
                info.entry.lead.name,
                xstr(info.entry.lead.source_name), xstr(info.excerpt),
                info.reliability.name, info.severity.name,
            ])

            # Column Name `Demographic Groups` Renamed to
            # `Vulnerable Group` as specified in Issue #280
            rows.permute_and_add(info.vulnerable_groups.all())
            rows.permute_and_add(info.specific_needs_groups.all())
            rows.permute_and_add(info.affected_groups.all())

            attributes = []
            if info.informationattribute_set.count() > 0:
                for attr in info.informationattribute_set.all():
                    attr_data = [attr.subpillar.pillar.name,
                                 attr.subpillar.name]

                    if attr.sector:
                        attr_data.append(attr.sector.name)
                        if attr.subsectors.count() > 0:
                            for ss in attr.subsectors.all():
                                attributes.append(attr_data + [ss.name])
                        else:
                            attributes.append(attr_data + [''])
                    else:
                        attributes.append(attr_data + ['', ''])
            else:
                attributes.append(['', '', '', ''])

            rows.permute_and_add_list(attributes)

            for country in countries:
                admin_levels = country.adminlevel_set.all()
                for admin_level in admin_levels:
                    selections = []
                    for map_selection in info.map_selections.all():
                        if admin_level == map_selection.admin_level:
                            selections.append(map_selection.name)
                    rows.permute_and_add(selections)

            ew.append(rows.rows, ws)
            grouped_rows.append(rows.group_rows)
        except:
            pass

    ew.append(grouped_rows, wsg)

    return ew.get_http_response(title)


def export_analysis_xls(title, event_pk=None, information_pks=None,
                        request_data=None):

    # Create a spreadsheet and get active workbook
    ew = ExcelWriter()
    ws = ew.get_active()
    ws.title = "Split Entries"
    wsg = ew.wb.create_sheet("Grouped Entries")

    # Create title row
    titles = [
        "Date of Lead Publication", "Imported By",
        "Date Imported", "URL", "Lead Title", "Source", "Excerpt"
    ]

    event = entry_models.Event.objects.get(pk=event_pk)
    elements = json.loads(event.entry_template.elements)
    sTypes = ['date-input', 'scale', 'number-input', 'multiselect',
              'organigram']
    element_ids = []
    geo_elements = []

    for element in elements:
        eType = element['type']
        if eType in sTypes:
            titles.append(element['label'])
        elif eType == 'matrix1d':
            titles.append([element['title'], 'Dimension', 'Sub-Dimension'])
        elif eType == 'matrix2d':
            titles.append([element['title'], 'Dimension', 'Sub-Dimension',
                           'Sector', 'Subsector'])
        elif eType == 'number2d':
            for row in element['rows']:
                for column in element['columns']:
                    titles.append('{} - {}'.format(row['title'],
                                                   column['title']))
            for row in element['rows']:
                titles.append('{} matches'.format(row['title']))
        elif eType == 'geolocations':
            geo_elements.append(element['id'])
        else:
            continue
        element_ids.append([element['id'], eType])

    if event_pk:
        countries = entry_models.Event.objects.get(pk=event_pk).countries.\
                                 all().distinct()
    else:
        countries = entry_models.Country.objects.all().distinct()

    for country in countries:
        admin_levels = country.adminlevel_set.all()
        for admin_level in admin_levels:
            titles.append(admin_level.name)
            titles.append('{} P-Code'.format(admin_level.name))

    index = 0
    for t in titles:
        if isinstance(t, list):
            for wswsg in [ws, wsg]:
                wswsg.cell(row=1, column=index+1).value = t[0]
                wswsg.cell(row=1, column=index+1).font = Font(bold=True)
                wswsg.merge_cells(start_row=1, end_row=1, start_column=index+1,
                                  end_column=index+len(t)-1)
            for ele in t[1:]:
                for wswsg in [ws, wsg]:
                    wswsg.cell(row=2, column=index+1).value = ele
                    wswsg.cell(row=2, column=index+1).font = Font(bold=True)
                index = index + 1
        else:
            for wswsg in [ws, wsg]:
                wswsg.cell(row=1, column=index+1).value = t
                wswsg.cell(row=1, column=index+1).font = Font(bold=True)
            index = index + 1

    ew.auto_fit_cells_in_row(1, ws)
    ew.auto_fit_cells_in_row(1, wsg)

    # Add each information in each entry belonging to this event
    informations = entry_models.EntryInformation.objects.filter(
        ~Q(entry__template=None),
        entry__lead__event__pk=event_pk).distinct()

    if information_pks:
        informations = informations.filter(pk__in=information_pks)
    informations = analysis_filter(informations, request_data, elements)

    grouped_rows = []
    for i, info in enumerate(informations):
        try:
            rows = RowCollection(1)

            lead_url = info.entry.lead.url
            if Attachment.objects.filter(lead=info.entry.lead).count() > 0:
                lead_url = info.entry.lead.attachment.upload.url

            rows.add_values([
                format_date(info.entry.lead.published_at),
                info.entry.created_by,
                format_date(info.entry.created_at.date()),
                lead_url,
                info.entry.lead.name, info.entry.lead.source_name,
                xstr(info.excerpt)
            ])

            infoE = json.loads(info.elements)
            for element_id, element_type in element_ids:
                element = list_filter(infoE, 'id', element_id)
                get_analysis_data(elements, element_id, element,
                                  element_type, rows)

            for country in countries:
                admin_levels = country.adminlevel_set.all()
                for admin_level in admin_levels:
                    selections = []
                    for map_selections in [geoE for geoE in infoE
                                           if geoE.get('id') in geo_elements]:
                        for map_selection in map_selections.get('value', []):
                            map_selection_list = map_selection.split(':')

                            if len(map_selection_list) == 3:
                                map_selection_list.append('')

                            if len(map_selection_list) == 4:
                                m_iso3, m_admin, m_name, pcode = \
                                    map_selection_list
                                if admin_level.level == int(m_admin):
                                    selections.append([m_name, pcode])

                    if len(selections) == 0:
                        selections = [['', '']]
                    rows.permute_and_add_list(selections)

            ew.append(rows.rows, ws)
            grouped_rows.append(rows.group_rows)
        except:
            pass

    ew.append(grouped_rows, wsg)

    # ew.save_to('/tmp/text.xls')  # REMOVE THIS
    return ew.get_http_response(title)


def export_and_save(event_pk, filename):
    # Create a spreadsheet and get active workbook
    ew = ExcelWriter()
    ws = ew.get_active()
    ws.title = "Split Entries"
    wsg = ew.wb.create_sheet("Grouped Entries")

    # Create title row
    titles = [
        "Country", "Date of Lead Publication", "Date of Information",
        "Created By",
        "Date Imported", "Lead Title", "Source", "Excerpt", "Reliability",
        "Severity", "Demographic Groups", "Specific Needs Groups",
        "Affected Groups", "Pillar", "Subpillar", "Sector", "Subsector",
    ]

    countries = entry_models.Event.objects.get(pk=event_pk).countries\
                            .all().distinct()

    for country in countries:
        admin_levels = country.adminlevel_set.all()
        for i, admin_level in enumerate(admin_levels):
            titles.append('Admin {}'.format(i))

    for i, t in enumerate(titles):
        ws.cell(row=1, column=i+1).value = t
        ws.cell(row=1, column=i+1).font = Font(bold=True)

        wsg.cell(row=1, column=i+1).value = t
        wsg.cell(row=1, column=i+1).font = Font(bold=True)

    ew.auto_fit_cells_in_row(1, ws)
    ew.auto_fit_cells_in_row(1, wsg)

    # Add each information in each entry belonging to this event
    informations = entry_models.EntryInformation.objects.filter(
                        entry__lead__event__pk=event_pk,
                        entry__template=None).distinct()

    grouped_rows = []
    for i, info in enumerate(informations):
        try:
            rows = RowCollection(1)

            rows.permute_and_add(info.entry.lead.event.countries.all())

            rows.add_values([
                format_date(info.entry.lead.published_at),
                format_date(info.date), info.entry.created_by,
                format_date(info.entry.created_at.date()),
                info.entry.lead.name, info.entry.lead.source_name,
                xstr(info.excerpt), info.reliability.name,
                info.severity.name
            ])

            # Column Name `Demographic Groups` Renamed to
            # `Vulnerable Group` as specified in Issue #280
            rows.permute_and_add(info.vulnerable_groups.all())
            rows.permute_and_add(info.specific_needs_groups.all())
            rows.permute_and_add(info.affected_groups.all())

            attributes = []
            if info.informationattribute_set.count() > 0:
                for attr in info.informationattribute_set.all():
                    attr_data = [attr.subpillar.pillar.name,
                                 attr.subpillar.name]

                    if attr.sector:
                        attr_data.append(attr.sector.name)
                        if attr.subsectors.count() > 0:
                            for ss in attr.subsectors.all():
                                attributes.append(attr_data + [ss.name])
                        else:
                            attributes.append(attr_data + [''])
                    else:
                        attributes.append(attr_data + ['', ''])
            else:
                attributes.append(['', '', '', ''])

            rows.permute_and_add_list(attributes)

            for country in countries:
                admin_levels = country.adminlevel_set.all()
                for admin_level in admin_levels:
                    selections = []
                    for map_selection in info.map_selections.all():
                        if admin_level == map_selection.admin_level:
                            selections.append(map_selection.name)
                    rows.permute_and_add(selections)

            ew.append(rows.rows, ws)
            grouped_rows.append(rows.group_rows)
        except:
            pass

    ew.append(grouped_rows, wsg)
    ew.save_to(filename)
